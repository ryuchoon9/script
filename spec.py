import openpyxl

# 데이터 추출
data = []

with open("spec.data", "r") as file:
    for line in file:
        data.append(line.strip())

hostname = data[0].split(":")[1].strip()
cpu_info = data[1].split(":")[1].strip()
memory_info = data[2].split(":")[1].strip()
disk_info = data[3].split(":")[1].strip()
release = data[4].split(":")[1].strip()
public_ip = data[5].split(":")[1].strip()
lb_ip = data[6].split(":")[1].strip()

# Excel 파일 생성
workbook = openpyxl.Workbook()
sheet = workbook.active

# 헤더 추가
sheet["A1"] = "Hostname"
sheet["B1"] = "CPU info"
sheet["C1"] = "Memory info"
sheet["D1"] = "Disk info"
sheet["E1"] = "Release"
sheet["F1"] = "Public IP"
sheet["G1"] = "LoadBalancer IP"

# 마지막 행 찾기
last_row = sheet.max_row + 1

# 데이터 추가
sheet.cell(row=last_row, column=1).value = hostname
sheet.cell(row=last_row, column=2).value = cpu_info
sheet.cell(row=last_row, column=3).value = memory_info
sheet.cell(row=last_row, column=4).value = disk_info
sheet.cell(row=last_row, column=5).value = release
sheet.cell(row=last_row, column=6).value = public_ip
sheet.cell(row=last_row, column=7).value = lb_ip

# Excel 파일 저장
workbook.save("server_specs.xlsx")
