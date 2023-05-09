import openpyxl

# 데이터 추출
data = []

with open("spec.data", "r") as file:
    for line in file:
        data.append(line.strip())

hostname = data[0].split(":")[1].strip()
cpu_info = data[1].split(":")[1].strip()
memory_info = data[2].split(":")[1].strip()
disk_info = data[3].split(":")[1].strip()
release = data[4].split(":")[1].strip()
public_ip = data[5].split(":")[1].strip()
lb_ip = data[6].split(":")[1].strip()

# Excel 파일 생성
workbook = openpyxl.load_workbook("server_specs.xlsx")
sheet = workbook.active

# 헤더 추가
sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
sheet["B1"] = "H/W"
sheet["A2"] = "Hostname"
sheet["B2"] = "CPU info"
sheet["C2"] = "Memory info"
sheet["D2"] = "Disk info"
sheet.merge_cells(start_row=1, start_column=5, end_row=1, end_column=8)
sheet["E1"] = "OS"
sheet["E2"] = "CentOS"
sheet["F2"] = "RHEL"
sheet["G2"] = "Ubuntu"
sheet["H2"] = "Windows"
sheet.merge_cells(start_row=1, start_column=9, end_row=1, end_column=10)
sheet["I1"] = "IP"
sheet["I2"] = "Public IP"
sheet["J2"] = "LoadBalancer IP"

# 마지막 행 찾기
last_row = sheet.max_row + 1

# 데이터 추가
sheet["A3"] = hostname
sheet["B3"] = cpu_info
sheet["C3"] = memory_info
sheet["D3"] = disk_info

# release 값을 소문자로 변환하여 검사
release_lower = release.lower()

if "centos" in release_lower:
    sheet["E3"] = release
elif "rhel" in release_lower:
    sheet["F3"] = release
elif "ubuntu" in release_lower:
    sheet["G3"] = release
elif "windows" in release_lower:
    sheet["H3"] = release

sheet["I3"] = public_ip
sheet["J3"] = lb_ip

# Excel 파일 저장
workbook.save("server_specs.xlsx")
